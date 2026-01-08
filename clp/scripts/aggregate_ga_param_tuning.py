
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# CONFIG: change these two if needed
# -----------------------------------------------------------------------------
BASE_RESULTS_GA_TUNE = Path(r"C:\Users\elahi\Desktop\clp\clp\results\ga_param_tuning")
OUT_CSV = BASE_RESULTS_GA_TUNE / "ga_param_tuning_summary.csv"
OUT_XLSX = BASE_RESULTS_GA_TUNE / "ga_param_tuning_summary.xlsx"


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _safe_get(d: Dict[str, Any], path: Tuple[str, ...], default: Any = None) -> Any:
    cur: Any = d
    for k in path:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur


def _fmt_float(x: Any) -> Any:
    try:
        return float(x)
    except Exception:
        return x


def _flatten_one(rec: Dict[str, Any], source_file: Path) -> Dict[str, Any]:
    """
    Convert one tuning JSON (one GA config aggregated over 10 cases)
    into one flat row for CSV/XLSX.
    """
    cfg = rec.get("ga_cfg", {}) or {}

    def stats(prefix: str) -> Dict[str, Any]:
        s = rec.get(prefix, {}) or {}
        return {
            f"{prefix}_A": _fmt_float(s.get("A", 0.0)),
            f"{prefix}_M1": _fmt_float(s.get("M1", 0.0)),
            f"{prefix}_M2": _fmt_float(s.get("M2", 0.0)),
        }

    row: Dict[str, Any] = {
        "br_class": rec.get("br_class", ""),
        "mode": rec.get("mode", ""),
        "seed": rec.get("seed", ""),
        "cases": ",".join(str(x) for x in rec.get("cases", [])),
        "Cr": _fmt_float(cfg.get("Cr", "")),
        "pm1": _fmt_float(cfg.get("pm1", "")),
        "pm2": _fmt_float(cfg.get("pm2", "")),
        "pop_size": cfg.get("pop_size", ""),
        "G": cfg.get("generations", ""),
        "source_file": str(source_file),
    }

    row.update(stats("Z1"))
    row.update(stats("Z2"))
    row.update(stats("Z3"))
    row.update(stats("rdx"))
    row.update(stats("rdy"))
    row.update(stats("rdz"))
    return row


def _write_csv(rows: List[Dict[str, Any]], out_path: Path, headers: List[str]) -> None:
    # Simple CSV writer (no pandas)
    import csv

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in headers})


def _autosize_columns(ws) -> None:
    # Basic autosize based on string length (good enough)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)


def _write_xlsx(rows: List[Dict[str, Any]], out_path: Path, headers: List[str]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "GA_Param_Tuning"

    # Header row
    ws.append(headers)

    # Data rows
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    _autosize_columns(ws)
    wb.save(out_path)


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def aggregate_ga_param_tuning() -> None:
    if not BASE_RESULTS_GA_TUNE.exists():
        raise FileNotFoundError(f"Tuning results folder not found: {BASE_RESULTS_GA_TUNE}")

    json_files = sorted(BASE_RESULTS_GA_TUNE.rglob("*.json"))
    rows: List[Dict[str, Any]] = []

    for fp in json_files:
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            continue

        if data.get("kind") != "ga_param_tuning_summary":
            continue

        rows.append(_flatten_one(data, fp))

    if not rows:
        print(f"No tuning summary JSONs found under: {BASE_RESULTS_GA_TUNE}")
        return

    # Stable header order (matches your table layout)
    headers = [
        "br_class", "mode", "seed", "cases",
        "Cr", "pm1", "pm2", "pop_size", "G",
        "Z1_A", "Z1_M1", "Z1_M2",
        "Z2_A", "Z2_M1", "Z2_M2",
        "Z3_A", "Z3_M1", "Z3_M2",
        "rdx_A", "rdx_M1", "rdx_M2",
        "rdy_A", "rdy_M1", "rdy_M2",
        "rdz_A", "rdz_M1", "rdz_M2",
        "source_file",
    ]

    # Sort rows for sanity: BR, mode, Cr, pm1, pm2, pop, G
    def sort_key(r: Dict[str, Any]) -> Tuple:
        return (
            str(r.get("br_class", "")),
            str(r.get("mode", "")),
            float(r.get("Cr", 0.0) or 0.0),
            float(r.get("pm1", 0.0) or 0.0),
            float(r.get("pm2", 0.0) or 0.0),
            int(r.get("pop_size", 0) or 0),
            int(r.get("G", 0) or 0),
        )

    rows.sort(key=sort_key)

    _write_csv(rows, OUT_CSV, headers)
    _write_xlsx(rows, OUT_XLSX, headers)

    print(f"✅ Wrote CSV:  {OUT_CSV}")
    print(f"✅ Wrote XLSX: {OUT_XLSX}")
    print(f"Rows: {len(rows)}")


