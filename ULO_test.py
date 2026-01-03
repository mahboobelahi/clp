import json
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

# Map box type (sorted dimensions) -> delivery rank pi
# Smaller pi = unloaded earlier
TYPE_TO_PI = {
    tuple(sorted((110, 43, 25))): 1,  # customer 3 (first)
    tuple(sorted((108, 76, 30))): 2,  # customer 2
    tuple(sorted((92, 81, 55))): 3,   # customer 1 (last)
}

# Boundary contact treated as NON-overlap (matches MILP)
TOUCH_IS_OVERLAP = False


# ============================================================
# DATA STRUCTURE
# ============================================================

@dataclass(frozen=True)
class Box:
    idx: int
    x: float
    y: float
    z: float
    l: float
    w: float
    h: float
    pi: int


# ============================================================
# GEOMETRY HELPERS
# ============================================================

def overlap_1d(a0, a1, b0, b1, *, touch_is_overlap=False):
    if touch_is_overlap:
        return not (a1 < b0 or b1 < a0)
    return (a0 < b1) and (b0 < a1)


def overlap_xy(i: Box, k: Box):
    return overlap_1d(i.x, i.x + i.l, k.x, k.x + k.l, touch_is_overlap=TOUCH_IS_OVERLAP) and \
           overlap_1d(i.y, i.y + i.w, k.y, k.y + k.w, touch_is_overlap=TOUCH_IS_OVERLAP)


def overlap_yz(i: Box, k: Box):
    return overlap_1d(i.y, i.y + i.w, k.y, k.y + k.w, touch_is_overlap=TOUCH_IS_OVERLAP) and \
           overlap_1d(i.z, i.z + i.h, k.z, k.z + k.h, touch_is_overlap=TOUCH_IS_OVERLAP)


def in_front_from_door_xL(i: Box, k: Box):
    # Door is rear face x = L → larger x is closer to door
    return k.x >= i.x + i.l


def above(i: Box, k: Box):
    return k.z >= i.z + i.h


# ============================================================
# ULO LOGIC (FINAL VERSION)
# ============================================================

def is_ulo(i: Box, k: Box) -> bool:
    if not (i.pi < k.pi):
        return False

    front_block = in_front_from_door_xL(i, k) and overlap_yz(i, k)
    above_block = above(i, k) and overlap_xy(i, k)

    return front_block or above_block


def compute_ulo_count(boxes: List[Box]) -> int:
    count = 0
    for i in boxes:
        for k in boxes:
            if i.idx != k.idx and is_ulo(i, k):
                count += 1
    return count


# ============================================================
# JSON PARSING
# ============================================================

def dims_key(box):
    return tuple(sorted((
        int(round(box["l"])),
        int(round(box["w"])),
        int(round(box["h"]))
    )))


def parse_boxes(solution) -> List[Box]:
    out = []
    for idx, b in enumerate(solution["layout"]["boxes"]):
        key = dims_key(b)
        if key not in TYPE_TO_PI:
            raise ValueError(f"Unknown box type: {key}")
        out.append(Box(
            idx=idx,
            x=b["x"], y=b["y"], z=b["z"],
            l=b["l"], w=b["w"], h=b["h"],
            pi=TYPE_TO_PI[key],
        ))
    return out


def list_json_files(path: Path) -> List[Path]:
    if path.is_file():
        return [path]
    return sorted(path.rglob("*.json"))


def parse_instance_meta(data):
    ds = data.get("dataset", {})
    instance_id = ds.get("instance_id", "")
    br_class = instance_id.split("_")[0] if "_" in instance_id else ""
    return ds.get("family", ""), instance_id, br_class


# ============================================================
# STATISTICS
# ============================================================

def safe_mode(vals):
    if not vals:
        return None
    freq = {}
    for v in vals:
        freq[v] = freq.get(v, 0) + 1
    return min(k for k, f in freq.items() if f == max(freq.values()))


def metrics(vals):
    if not vals:
        return dict(mean=None, median=None, mode=None, min=None, max=None, std=None)
    return dict(
        mean=float(statistics.mean(vals)),
        median=float(statistics.median(vals)),
        mode=float(safe_mode(vals)),
        min=float(min(vals)),
        max=float(max(vals)),
        std=float(statistics.pstdev(vals)) if len(vals) > 1 else 0.0,
    )


# ============================================================
# XLSX HELPERS
# ============================================================

def autosize(ws):
    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, width + 2)


def write_xlsx_flat(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ulo_flat"

    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])

    autosize(ws)
    wb.save(path)


def write_xlsx_summary(file_rows, class_rows, path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "per_file"
    headers = list(file_rows[0].keys()) if file_rows else []
    ws1.append(headers)
    for r in file_rows:
        ws1.append([r[h] for h in headers])
    autosize(ws1)

    ws2 = wb.create_sheet("per_br_class")
    headers = list(class_rows[0].keys()) if class_rows else []
    ws2.append(headers)
    for r in class_rows:
        ws2.append([r[h] for h in headers])
    autosize(ws2)

    wb.save(path)


# ============================================================
# MAIN ENTRY FUNCTION (THIS IS WHAT YOU CALL)
# ============================================================

def run_ulo_analysis(input_path, out_flat="ulo_flat.xlsx", 
                     out_summary="ulo_summary.xlsx"):
    
    files = list_json_files(Path(input_path))
    if not files:
        raise FileNotFoundError("No JSON files found")

    flat_rows = []
    file_rows = []
    class_map = {}

    for fp in files:
        data = json.loads(fp.read_text())
        family, instance_id, br_class = parse_instance_meta(data)

        ulo_vals = []
        for sol in data.get("pareto_front", []):
            boxes = parse_boxes(sol)
            ulo = compute_ulo_count(boxes)
            ulo_vals.append(ulo)

            flat_rows.append(dict(
                file=str(fp),
                family=family,
                instance_id=instance_id,
                br_class=br_class,
                solution_id=sol.get("solution_id", ""),
                ulo_count=ulo,
                n_boxes=len(boxes),
            ))

        m = metrics(ulo_vals)
        file_rows.append(dict(
            file=str(fp),
            family=family,
            instance_id=instance_id,
            br_class=br_class,
            n_solutions=len(ulo_vals),
            **{f"{k}_ulo": v for k, v in m.items()},
        ))

        class_map.setdefault(br_class, []).extend(ulo_vals)

    class_rows = []
    for br, vals in class_map.items():
        m = metrics(vals)
        class_rows.append(dict(
            br_class=br,
            n_solutions=len(vals),
            **{f"{k}_ulo": v for k, v in m.items()},
        ))

    write_xlsx_flat(flat_rows, out_flat)
    write_xlsx_summary(file_rows, class_rows, out_summary)

    return {
        "files": len(files),
        "solutions": len(flat_rows),
        "flat_xlsx": out_flat,
        "summary_xlsx": out_summary,
    }
# ============================================================
if __name__ == "__main__":
    run_ulo_analysis()